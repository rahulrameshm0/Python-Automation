import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

wb = load_workbook('pivot_table.xlsx')
sheet = wb['Report']

min_column = wb.active.min_column
max_column = wb.active.max_column
min_row = wb.active.min_row
max_row = wb.active.max_row

# sheet['B8'] = '=SUM(B6:B7)'
# sheet['B8'].style = 'Currency'
#
# sheet['C8'] = '=SUM(C6:C7)'
# sheet['C8'].style = 'Currency'
#
# sheet['D8'] = '=SUM(D6:D7)'
# sheet['D8'].style = 'Currency'
#
# sheet['E8'] = '=SUM(E6:E7)'
# sheet['E8'].style = 'Currency'
#
# sheet['F8'] = '=SUM(F6:F7)'
# sheet['F8'].style = 'Currency'
#
# sheet['G8'] = '=SUM(G6:G7)'
# sheet['G8'].style = 'Currency'

for i in range(min_column+1, max_column+1):
    letter = get_column_letter(i)
    sheet[f'{letter}{max_row+1}'] = f'=SUM({letter}{min_row+1}:{letter}{max_row})'
    sheet[f'{letter}{max_row+1}'].style = 'Currency'

wb.save('report.xlsx')